#!/usr/bin/env python3
"""
cashfree_processor.py
End-to-end Cashfree Transfer Report processor.
Implements all 10 steps from the cashfree-skill.

Usage:
    python3 cashfree_processor.py <input_xlsx> <output_xlsx>

Progress is emitted to stdout as JSON lines:
    {"step": 1, "total": 10, "label": "...", "detail": "..."}
    {"step": 10, "total": 10, "label": "Complete", "detail": "...", "done": true}
    {"error": "..."}  (on fatal error)
"""

import sys
import json
import re
import datetime
import traceback

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ─── Style constants ──────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
SUB_FILL  = PatternFill("solid", fgColor="2E75B6")
TLY_HDR   = PatternFill("solid", fgColor="375623")
DARK_ORG  = PatternFill("solid", fgColor="843C0C")
NEW_FILL  = PatternFill("solid", fgColor="E2EFDA")
ORANGE    = PatternFill("solid", fgColor="FFC000")
TOT_FILL  = PatternFill("solid", fgColor="BDD7EE")
INFO_FILL = PatternFill("solid", fgColor="F2F2F2")
ALT_FILL  = PatternFill("solid", fgColor="DEEAF1")
YELL_FILL = PatternFill("solid", fgColor="FFFF00")

HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
SUB_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
TLY_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
TOT_FONT  = Font(bold=True, name="Arial", size=10)
DAT_FONT  = Font(name="Arial", size=10)
BOLD_FONT = Font(bold=True, name="Arial", size=10)

CTR_WRAP  = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT_ALG = Alignment(horizontal='right')
LEFT_ALG  = Alignment(horizontal='left')
CTR_ALG   = Alignment(horizontal='center')

def emit(step, total, label, detail="", done=False, extra=None):
    obj = {"step": step, "total": total, "label": label, "detail": detail}
    if done:
        obj["done"] = True
    if extra:
        obj.update(extra)
    print(json.dumps(obj), flush=True)

def emit_error(msg):
    print(json.dumps({"error": msg}), flush=True)

# ─── Helper cell writers ──────────────────────────────────────────────────────
def H(ws, r, c, v, fill=None, font=None):
    cell = ws.cell(r, c, v)
    cell.fill = fill or HDR_FILL
    cell.font = font or HDR_FONT
    cell.alignment = CTR_WRAP
    return cell

def D(ws, r, c, v, bold=False, fill=None):
    cell = ws.cell(r, c, v)
    cell.font = Font(bold=bold, name="Arial", size=10)
    if fill:
        cell.fill = fill
    return cell

def N(ws, r, c, v, fmt='#,##0.00', bold=False, fill=None):
    cell = ws.cell(r, c, v)
    cell.font = Font(bold=bold, name="Arial", size=10)
    cell.number_format = fmt
    cell.alignment = RIGHT_ALG
    if fill:
        cell.fill = fill
    return cell

# ─── Week helper ──────────────────────────────────────────────────────────────
def week_end_sunday(d):
    days = (6 - d.weekday()) % 7
    return d + datetime.timedelta(days=days)


def process(src_path, out_path):
    TOTAL = 10

    # ── Pre-check ────────────────────────────────────────────────────────────
    emit(0, TOTAL, "Reading file", "Loading workbook sheets…")
    try:
        xl = pd.read_excel(src_path, sheet_name=None, dtype=str)
    except Exception as e:
        emit_error(f"Cannot read file: {e}")
        return

    sheets = list(xl.keys())
    tr_key = next((s for s in sheets if 'transfer' in s.lower()), None)
    ac_key = next((s for s in sheets if 'account' in s.lower()), None)

    if not tr_key:
        emit_error("Missing 'transfer report' sheet in the uploaded file.")
        return
    if not ac_key:
        emit_error("Missing 'account statement' sheet in the uploaded file.")
        return

    tr   = xl[tr_key].copy()
    acct = xl[ac_key].copy()

    # ── Normalise column names ────────────────────────────────────────────────
    # Cashfree sometimes truncates 'Transfer Id' to 'Tra' in the export
    tr.rename(columns={'Tra': 'Transfer Id'}, inplace=True)
    # Drop any pre-existing countif / count columns — we regenerate them
    for col in ['countif', 'count', 'cc']:
        if col in tr.columns:
            tr.drop(columns=[col], inplace=True)
    for col in ['count', 'cc']:
        if col in acct.columns:
            acct.drop(columns=[col], inplace=True)

    # ── STEP 1 — Filter rows ─────────────────────────────────────────────────────
    # Rules:
    #   1. FAILED / REJECTED  → always remove
    #   2. SUCCESS / RECEIVED → remove if Service Charge == 0 OR Service Tax == 0 / blank
    emit(1, TOTAL, "Cleaning rows",
         "Removing FAILED/REJECTED rows; also removing SUCCESS/RECEIVED rows with zero Service Charge or Service Tax…")
    try:
        mask_fr      = tr['Status'].isin(['FAILED', 'REJECTED'])
        # Treat blank/non-numeric as 0
        sc_numeric   = pd.to_numeric(tr['Service Charge'], errors='coerce').fillna(0)
        st_numeric   = pd.to_numeric(tr['Service Tax'],    errors='coerce').fillna(0)
        mask_zero_sc = sc_numeric == 0
        mask_zero_st = st_numeric == 0
        # For non-failure rows, remove when Service Charge OR Service Tax is zero/blank
        mask_zero_charges = mask_zero_sc | mask_zero_st
        to_remove    = mask_fr | (~mask_fr & mask_zero_charges)
        removed      = int(to_remove.sum())
        removed_fr   = int(mask_fr.sum())
        removed_zc   = int((~mask_fr & mask_zero_charges).sum())
        tr_clean     = tr[~to_remove].copy().reset_index(drop=True)
        status_bd    = tr_clean['Status'].value_counts().to_dict()
        emit(1, TOTAL, "Step 1 complete",
             f"Removed {removed} rows total: {removed_fr} FAILED/REJECTED, "
             f"{removed_zc} SUCCESS/RECEIVED with zero Service Charge or Service Tax. "
             f"Remaining: {len(tr_clean)}. Status breakdown: {status_bd}",
             extra={"removed": removed, "remaining": len(tr_clean)})
    except Exception as e:
        emit_error(f"Step 1 failed: {e}")
        return

    # ── STEP 2 — Normalise Transfer IDs ───────────────────────────────────
    emit(2, TOTAL, "Normalising Transfer IDs", "Stripping n-suffixes from Transfer IDs…")
    try:
        def clean_tid(tid):
            return re.sub(r'n\d+$', '', str(tid).strip())

        original_ids = tr_clean['Transfer Id'].copy()
        tr_clean['Transfer Id'] = tr_clean['Transfer Id'].apply(clean_tid)
        suffix_fixed = int((original_ids != tr_clean['Transfer Id']).sum())
        examples_before = list(original_ids[original_ids != tr_clean['Transfer Id']].head(3))
        examples_after  = list(tr_clean.loc[original_ids != tr_clean['Transfer Id'], 'Transfer Id'].head(3))
        emit(2, TOTAL, "Step 2 complete",
             f"Stripped n-suffixes from {suffix_fixed} Transfer IDs. Examples: {examples_before} → {examples_after}",
             extra={"suffix_fixed": suffix_fixed})
    except Exception as e:
        emit_error(f"Step 2 failed: {e}")
        return

    # ── STEP 3 — Duplicate check ───────────────────────────────────────────
    emit(3, TOTAL, "Checking for duplicate Transfer IDs", "Scanning for duplicates…")
    try:
        dup_mask  = tr_clean['Transfer Id'].duplicated(keep=False)
        dup_count = int(dup_mask.sum())
        dup_ids   = tr_clean.loc[dup_mask, 'Transfer Id'].unique().tolist()
        emit(3, TOTAL, "Step 3 complete",
             f"Duplicate rows: {dup_count} across {len(dup_ids)} unique Transfer IDs. {'Highlighted in yellow in output.' if dup_count > 0 else 'No duplicates.'}",
             extra={"dup_count": dup_count, "dup_ids_count": len(dup_ids)})
    except Exception as e:
        emit_error(f"Step 3 failed: {e}")
        return

    # ── STEP 4 — Subtotal reconciliation ──────────────────────────────────
    emit(4, TOTAL, "Subtotal reconciliation", "Comparing transfer report vs account statement totals…")
    try:
        tr_success = tr_clean[tr_clean['Status'] == 'SUCCESS']
        acct_pt    = acct[acct['Particulars'] == 'PAYOUT_TRANSFER']
        recon_lines = []
        for col_tr, col_ac in [('Amount', 'Amount (INR)'),
                                ('Service Charge', 'Service Charge (INR)'),
                                ('Service Tax', 'Service Tax (INR)')]:
            tr_val = round(tr_success[col_tr].astype(float).sum(), 2)
            ac_val = round(acct_pt[col_ac].astype(float).sum(), 2)
            status = "MATCH" if tr_val == ac_val else f"MISMATCH diff={tr_val - ac_val:.2f}"
            recon_lines.append(f"{col_tr}: TR={tr_val:,.2f} Acct={ac_val:,.2f} → {status}")
        emit(4, TOTAL, "Step 4 complete", " | ".join(recon_lines))
    except Exception as e:
        emit_error(f"Step 4 failed: {e}")
        return

    # ── STEP 5 — Write cleaned transfer report + count column ─────────────
    emit(5, TOTAL, "Writing transfer report sheet", "Adding count column and processing log…")
    try:
        wb = openpyxl.load_workbook(src_path)
        # Remove existing sheets we'll recreate
        for sname in ['transfer report', 'account statement', 'Summary', 'tally entry']:
            if sname in wb.sheetnames:
                del wb[sname]

        ws_tr = wb.create_sheet('transfer report')

        # Processing log rows 1-4
        ws_tr.cell(1, 1, "PROCESSING LOG").font = Font(bold=True, name="Arial", size=11, color="1F4E79")
        ws_tr.cell(2, 1, f"Step 1: Removed {removed} rows ({removed_fr} FAILED/REJECTED, {removed_zc} SUCCESS/RECEIVED with zero Service Charge or Service Tax)").font = DAT_FONT
        ws_tr.cell(3, 1, f"Step 2: Fixed {suffix_fixed} Transfer IDs (removed nN suffixes)").font = DAT_FONT
        ws_tr.cell(4, 1, f"Step 3: Duplicates found: {dup_count} rows across {len(dup_ids)} unique IDs").font = DAT_FONT

        # Headers at row 5 — insert 'count' after Transfer Id (col C = index 2)
        orig_cols = list(tr_clean.columns)
        tid_idx   = orig_cols.index('Transfer Id')  # 0-based
        new_cols  = orig_cols[:tid_idx+1] + ['count'] + orig_cols[tid_idx+1:]

        for ci, col in enumerate(new_cols, 1):
            if col == 'count':
                H(ws_tr, 5, ci, 'count', fill=TLY_HDR, font=TLY_FONT)
            else:
                H(ws_tr, 5, ci, col)

        # Data rows starting at row 6
        count_col_excel = tid_idx + 2  # 1-based col of 'count' in Excel (C=3 → count=4 → D)
        count_col_letter = get_column_letter(count_col_excel)
        tid_col_letter   = get_column_letter(tid_idx + 1)  # C

        for ri, row in tr_clean.iterrows():
            excel_row = ri + 6  # row 6 = first data row
            col_offset = 0
            for ci, col in enumerate(new_cols, 1):
                if col == 'count':
                    # COUNTIF formula
                    r = excel_row
                    formula = f'=COUNTIF(${tid_col_letter}:${tid_col_letter},IFERROR(LEFT({tid_col_letter}{r},FIND("n",{tid_col_letter}{r},LEN({tid_col_letter}{r})-3)-1),{tid_col_letter}{r})&"*")'
                    cell = ws_tr.cell(excel_row, ci, formula)
                    cell.fill = NEW_FILL
                    cell.font = DAT_FONT
                    cell.alignment = CTR_ALG
                    cell.number_format = '0'
                else:
                    val = row[col]
                    cell = ws_tr.cell(excel_row, ci, val)
                    cell.font = DAT_FONT
                    # Yellow for duplicate Transfer IDs
                    if col == 'Transfer Id' and dup_mask.iloc[ri]:
                        cell.fill = YELL_FILL

        # Column widths
        ws_tr.column_dimensions['A'].width = 22
        ws_tr.column_dimensions['B'].width = 14
        ws_tr.column_dimensions[tid_col_letter].width = 25
        ws_tr.column_dimensions[count_col_letter].width = 8

        emit(5, TOTAL, "Step 5 complete",
             f"Transfer report written with {len(tr_clean)} rows + count column (col {count_col_letter}).")
    except Exception as e:
        emit_error(f"Step 5 failed: {e}\n{traceback.format_exc()}")
        return

    # ── STEP 6 — Enrich account statement ─────────────────────────────────
    emit(6, TOTAL, "Enriching account statement", "Adding cc (INDEX/MATCH) and count columns…")
    try:
        ws_ac = wb.create_sheet('account statement')

        acct_cols = list(acct.columns)
        eid_idx   = acct_cols.index('Event Id')  # 0-based
        new_ac_cols = acct_cols[:eid_idx+1] + ['cc', 'count'] + acct_cols[eid_idx+1:]

        # Headers
        for ci, col in enumerate(new_ac_cols, 1):
            if col in ('cc', 'count'):
                H(ws_ac, 1, ci, col, fill=TLY_HDR, font=TLY_FONT)
            else:
                H(ws_ac, 1, ci, col)

        # Data rows
        eid_col_excel = eid_idx + 1   # 1-based col of Event Id in the NEW layout
        cc_col_excel  = eid_idx + 2   # J
        cnt_col_excel = eid_idx + 3   # K
        eid_letter    = get_column_letter(eid_col_excel)
        cc_letter     = get_column_letter(cc_col_excel)

        matched = 0
        total_ac = len(acct)
        for ri, row in acct.iterrows():
            excel_row = ri + 2
            for ci, col in enumerate(new_ac_cols, 1):
                if col == 'cc':
                    r = excel_row
                    formula = f"=IFERROR(INDEX('transfer report'!$C:$C,MATCH({eid_letter}{r},'transfer report'!$E:$E,0)),\"\")"
                    cell = ws_ac.cell(excel_row, ci, formula)
                    cell.fill = NEW_FILL
                    cell.font = DAT_FONT
                    # Count matched (heuristic: non-blank Event Id)
                    if str(row.get('Event Id', '')).strip():
                        matched += 1
                elif col == 'count':
                    r = excel_row
                    formula = f'=IF({cc_letter}{r}="","",IFERROR(COUNTIF(\'transfer report\'!$C:$C,IFERROR(LEFT({cc_letter}{r},FIND("n",{cc_letter}{r},LEN({cc_letter}{r})-3)-1),{cc_letter}{r})&"*"),0))'
                    cell = ws_ac.cell(excel_row, ci, formula)
                    cell.fill = NEW_FILL
                    cell.font = DAT_FONT
                    cell.number_format = '0'
                else:
                    val = row[col]
                    cell = ws_ac.cell(excel_row, ci, val)
                    cell.font = DAT_FONT

        # Column widths
        ws_ac.column_dimensions['A'].width = 22
        ws_ac.column_dimensions[eid_letter].width = 14
        ws_ac.column_dimensions[cc_letter].width = 22
        ws_ac.column_dimensions[get_column_letter(cnt_col_excel)].width = 8

        emit(6, TOTAL, "Step 6 complete",
             f"Account statement enriched: cc (col {cc_letter}) + count (col {get_column_letter(cnt_col_excel)}). Rows: {total_ac}.")
    except Exception as e:
        emit_error(f"Step 6 failed: {e}\n{traceback.format_exc()}")
        return

    # ── STEP 7 — Build Summary sheet ──────────────────────────────────────
    emit(7, TOTAL, "Building Summary sheet", "Creating pivot tables and Trf Entries…")
    try:
        ws_sum = wb.create_sheet('Summary')
        OFF = 7  # PAYOUT table starts at col G (1-based = 7)

        # Parse dates
        acct_work = acct.copy()
        acct_work['date_only'] = pd.to_datetime(acct_work['Added On'], errors='coerce').dt.normalize()
        acct_work['Amount (INR)']          = pd.to_numeric(acct_work['Amount (INR)'], errors='coerce').fillna(0)
        acct_work['Service Charge (INR)']  = pd.to_numeric(acct_work['Service Charge (INR)'], errors='coerce').fillna(0)
        acct_work['Service Tax (INR)']     = pd.to_numeric(acct_work['Service Tax (INR)'], errors='coerce').fillna(0)

        # ── Section A — Pivot config legend (rows 1-8) ────────────────────
        ws_sum.merge_cells('A1:F1')
        c = ws_sum.cell(1, 1, "PIVOT TABLE CONFIGURATION — Account Statement")
        c.font = Font(bold=True, name="Arial", size=12, color="1F4E79")
        c.alignment = Alignment(horizontal='center')

        config_hdrs = ["Area", "Field", "Description"]
        for ci, h in enumerate(config_hdrs, 1):
            cell = ws_sum.cell(2, ci, h)
            cell.fill = SUB_FILL
            cell.font = SUB_FONT
        ws_sum.merge_cells('C2:F2')

        config_rows = [
            ("ROWS",    "Added On (date)",         "Groups data by transfer date"),
            ("VALUES",  "Sum of Amount (INR)",      "Total payout per day"),
            ("VALUES",  "Sum of Service Charge (INR)", "Total charges per day"),
            ("VALUES",  "Sum of Service Tax (INR)", "Total tax per day"),
            ("FILTERS", "Particulars",              "PAYOUT_TRANSFER / BANK_TRANSFER / VPA_DETAILS_FROM_VPA"),
            ("FILTERS", "count (of Event Id)",      "Filter by Event Id occurrence count"),
        ]
        for i, (area, field, desc) in enumerate(config_rows, 3):
            fill = ALT_FILL if i % 2 == 0 else INFO_FILL
            ws_sum.cell(i, 1, area).fill = fill
            ws_sum.cell(i, 2, field).fill = fill
            ws_sum.cell(i, 3, desc).fill = fill
            ws_sum.merge_cells(f'C{i}:F{i}')

        # ── Section B — Full pivot (all particulars), cols A-D ────────────
        ROW_START_B = 10
        # Filter labels
        ws_sum.cell(ROW_START_B, 1, "Particulars").font = BOLD_FONT
        ws_sum.cell(ROW_START_B, 2, "(All)").font = DAT_FONT
        ws_sum.cell(ROW_START_B + 1, 1, "count").font = BOLD_FONT
        ws_sum.cell(ROW_START_B + 1, 2, "(All)").font = DAT_FONT

        # Headers
        hdr_row_b = ROW_START_B + 2
        for ci, h in enumerate(["Row Labels", "Sum of Amount (INR)", "Sum of Service Charge (INR)", "Sum of Service Tax (INR)"], 1):
            H(ws_sum, hdr_row_b, ci, h)

        all_daily = acct_work.groupby('date_only').agg(
            amt=('Amount (INR)', 'sum'),
            sc=('Service Charge (INR)', 'sum'),
            st=('Service Tax (INR)', 'sum'),
        ).reset_index().sort_values('date_only')

        data_start_b = hdr_row_b + 1
        for ri, row in all_daily.iterrows():
            r = data_start_b + list(all_daily.index).index(ri)
            date_str = row['date_only'].strftime('%Y-%m-%d') if pd.notna(row['date_only']) else ''
            D(ws_sum, r, 1, date_str)
            N(ws_sum, r, 2, row['amt'])
            N(ws_sum, r, 3, row['sc'])
            N(ws_sum, r, 4, row['st'])

        data_end_b = data_start_b + len(all_daily) - 1
        gt_b = data_end_b + 1
        D(ws_sum, gt_b, 1, "Grand Total", bold=True, fill=TOT_FILL)
        for ci in range(2, 5):
            col_l = get_column_letter(ci)
            N(ws_sum, gt_b, ci, f"=SUM({col_l}{data_start_b}:{col_l}{data_end_b})", bold=True, fill=TOT_FILL)

        # ── Section C — PAYOUT_TRANSFER table, cols G onward ─────────────
        pt_data = acct_work[acct_work['Particulars'] == 'PAYOUT_TRANSFER'].copy()
        pt = pt_data.groupby('date_only').agg(
            amt=('Amount (INR)', 'sum'),
            sc=('Service Charge (INR)', 'sum'),
            st=('Service Tax (INR)', 'sum'),
        ).reset_index().sort_values('date_only')
        pt['net'] = pt['sc'] + pt['st']

        # Weekly Trf Entries
        pt['week_end'] = pt['date_only'].apply(lambda d: week_end_sunday(d.date()) if pd.notna(d) else None)
        trf_entries = {}
        week_ranges = {}
        for we, grp in pt.groupby('week_end'):
            net_sum    = round(float(grp['net'].sum()), 2)
            last_date  = grp['date_only'].max()
            start_date = grp['date_only'].min()
            trf_entries[last_date] = net_sum
            week_ranges[last_date] = (start_date, last_date)

        ROW_START_C = ROW_START_B
        ws_sum.cell(ROW_START_C, OFF, "Particulars").font = BOLD_FONT
        ws_sum.cell(ROW_START_C, OFF + 1, "PAYOUT_TRANSFER").font = DAT_FONT
        ws_sum.cell(ROW_START_C + 1, OFF, "count").font = BOLD_FONT
        ws_sum.cell(ROW_START_C + 1, OFF + 1, "(All)").font = DAT_FONT

        hdr_row_c = ROW_START_C + 2
        for ci, h in enumerate(["Row Labels", "Sum of Amount (INR)", "Sum of Service Charge (INR)",
                                  "Sum of Service Tax (INR)", "net", "Trf Entries"], OFF):
            if h in ("net",):
                H(ws_sum, hdr_row_c, ci, h, fill=TLY_HDR, font=TLY_FONT)
            elif h == "Trf Entries":
                H(ws_sum, hdr_row_c, ci, h, fill=PatternFill("solid", fgColor="843C0C"), font=HDR_FONT)
            else:
                H(ws_sum, hdr_row_c, ci, h)

        data_start_c = hdr_row_c + 1
        pt_data_start = data_start_c  # for tally entry references
        pt_row_map = {}  # date → excel row

        for ri, row in enumerate(pt.itertuples(), 0):
            r = data_start_c + ri
            date_str = row.date_only.strftime('%Y-%m-%d') if pd.notna(row.date_only) else ''
            D(ws_sum, r, OFF, date_str)
            N(ws_sum, r, OFF + 1, row.amt)
            N(ws_sum, r, OFF + 2, row.sc)
            N(ws_sum, r, OFF + 3, row.st)
            # net formula
            sc_l = get_column_letter(OFF + 2)
            st_l = get_column_letter(OFF + 3)
            net_cell = ws_sum.cell(r, OFF + 4, f"={sc_l}{r}+{st_l}{r}")
            net_cell.font = DAT_FONT
            net_cell.number_format = '#,##0.00'
            net_cell.fill = NEW_FILL
            # Trf Entries
            if row.date_only in trf_entries:
                trf_cell = ws_sum.cell(r, OFF + 5, trf_entries[row.date_only])
                trf_cell.fill = ORANGE
                trf_cell.font = BOLD_FONT
                trf_cell.number_format = '#,##0.00'
            pt_row_map[row.date_only] = r

        data_end_c = data_start_c + len(pt) - 1
        gt_c = data_end_c + 1
        D(ws_sum, gt_c, OFF, "Grand Total", bold=True, fill=TOT_FILL)
        for ci in range(OFF + 1, OFF + 6):
            col_l = get_column_letter(ci)
            N(ws_sum, gt_c, ci, f"=SUM({col_l}{data_start_c}:{col_l}{data_end_c})", bold=True, fill=TOT_FILL)

        # ── Section D — VPA_DETAILS_FROM_VPA table ────────────────────────
        vpa_data = acct_work[acct_work['Particulars'] == 'VPA_DETAILS_FROM_VPA'].copy()
        vpa = vpa_data.groupby('date_only').agg(
            amt=('Amount (INR)', 'sum'),
            sc=('Service Charge (INR)', 'sum'),
            st=('Service Tax (INR)', 'sum'),
        ).reset_index().sort_values('date_only')
        vpa['net'] = vpa['sc'] + vpa['st']

        ROW_START_D = gt_c + 3
        ws_sum.cell(ROW_START_D, OFF, "Particulars").font = BOLD_FONT
        ws_sum.cell(ROW_START_D, OFF + 1, "VPA_DETAILS_FROM_VPA").font = DAT_FONT
        ws_sum.cell(ROW_START_D + 1, OFF, "count").font = BOLD_FONT
        ws_sum.cell(ROW_START_D + 1, OFF + 1, "(All)").font = DAT_FONT

        hdr_row_d = ROW_START_D + 2
        for ci, h in enumerate(["Row Labels", "Sum of Amount (INR)", "Sum of Service Charge (INR)",
                                  "Sum of Service Tax (INR)", "net", "VPA_DETAILS_FROM_VPA"], OFF):
            if h in ("net",):
                H(ws_sum, hdr_row_d, ci, h, fill=TLY_HDR, font=TLY_FONT)
            elif h == "VPA_DETAILS_FROM_VPA":
                H(ws_sum, hdr_row_d, ci, h, fill=PatternFill("solid", fgColor="843C0C"), font=HDR_FONT)
            else:
                H(ws_sum, hdr_row_d, ci, h)

        data_start_d = hdr_row_d + 1
        vpa_row_map = {}
        for ri, row in enumerate(vpa.itertuples(), 0):
            r = data_start_d + ri
            date_str = row.date_only.strftime('%Y-%m-%d') if pd.notna(row.date_only) else ''
            D(ws_sum, r, OFF, date_str)
            N(ws_sum, r, OFF + 1, row.amt)
            N(ws_sum, r, OFF + 2, row.sc)
            N(ws_sum, r, OFF + 3, row.st)
            sc_l = get_column_letter(OFF + 2)
            st_l = get_column_letter(OFF + 3)
            net_cell = ws_sum.cell(r, OFF + 4, f"={sc_l}{r}+{st_l}{r}")
            net_cell.font = DAT_FONT
            net_cell.number_format = '#,##0.00'
            net_cell.fill = NEW_FILL
            # VPA_DETAILS col = net value
            vpa_cell = ws_sum.cell(r, OFF + 5, row.net)
            vpa_cell.fill = ORANGE
            vpa_cell.font = BOLD_FONT
            vpa_cell.number_format = '#,##0.00'
            vpa_row_map[row.date_only] = r

        if len(vpa) > 0:
            data_end_d = data_start_d + len(vpa) - 1
            gt_d = data_end_d + 1
            D(ws_sum, gt_d, OFF, "Grand Total", bold=True, fill=TOT_FILL)
            for ci in range(OFF + 1, OFF + 6):
                col_l = get_column_letter(ci)
                N(ws_sum, gt_d, ci, f"=SUM({col_l}{data_start_d}:{col_l}{data_end_d})", bold=True, fill=TOT_FILL)

        # Column widths for Summary
        col_widths = {1:14, 2:22, 3:28, 4:22, 5:6, 6:6, 7:14, 8:22, 9:28, 10:22, 11:14, 12:20}
        for ci, w in col_widths.items():
            ws_sum.column_dimensions[get_column_letter(ci)].width = w

        pt_net_total = float(pt['net'].sum())
        vpa_net_total = float(vpa['net'].sum()) if len(vpa) > 0 else 0.0
        emit(7, TOTAL, "Step 7 complete",
             f"Summary: {len(all_daily)} all-daily rows | PAYOUT: {len(pt)} rows, net={pt_net_total:.2f} | "
             f"Trf Entries: {len(trf_entries)} | VPA: {len(vpa)} rows, net={vpa_net_total:.2f}",
             extra={"pt_rows": len(pt), "vpa_rows": len(vpa), "trf_entries": len(trf_entries)})
    except Exception as e:
        emit_error(f"Step 7 failed: {e}\n{traceback.format_exc()}")
        return

    # ── STEP 8 — Build Tally Entry sheet ──────────────────────────────────
    emit(8, TOTAL, "Building tally entry sheet", "Creating journal entry rows…")
    try:
        ws_te = wb.create_sheet('tally entry')
        te_cols = ["entry_code", "DATE", "Tally Date", "Mode", "VOUCHERTYPENAME",
                   "NARRATION", "DebitLedger", "AmountDebitLedger", "CreditLedger", "AmountCreditLedger"]
        for ci, h in enumerate(te_cols, 1):
            H(ws_te, 1, ci, h, fill=TLY_HDR, font=TLY_FONT)

        AMT_COL = get_column_letter(OFF + 1)   # H
        TRF_COL = get_column_letter(OFF + 5)   # L

        tally_row = 2
        entry_code = 1

        # Part 1 — Normal rows (one per PAYOUT date)
        for ri, row in enumerate(pt.itertuples(), 0):
            r = tally_row
            sum_row = pt_data_start + ri
            date_str    = row.date_only.strftime('%d/%m/%y') if pd.notna(row.date_only) else ''
            tally_date  = int(row.date_only.strftime('%Y%m%d')) if pd.notna(row.date_only) else ''
            narration   = f"Refund Done through Cashfree  on {row.date_only.strftime('%Y-%m-%d')}" if pd.notna(row.date_only) else ''

            ws_te.cell(r, 1, entry_code).font = DAT_FONT
            ws_te.cell(r, 2, date_str).font = DAT_FONT
            ws_te.cell(r, 3, tally_date).font = DAT_FONT
            ws_te.cell(r, 3).number_format = '0'
            ws_te.cell(r, 4, "Journal").font = DAT_FONT
            ws_te.cell(r, 5, "P- JV- MH").font = DAT_FONT
            ws_te.cell(r, 6, narration).font = DAT_FONT
            ws_te.cell(r, 7, "Refund Pay to Customer-COD").font = DAT_FONT
            # AmountDebitLedger = -J{r}
            ws_te.cell(r, 8, f"=-J{r}").font = DAT_FONT
            ws_te.cell(r, 8).number_format = '#,##0.00'
            ws_te.cell(r, 9, "Cashfree Payments India(Pasfar) Pvt Ltd-C").font = DAT_FONT
            # AmountCreditLedger = Summary!H{sum_row}
            ws_te.cell(r, 10, f"=Summary!{AMT_COL}{sum_row}").font = DAT_FONT
            ws_te.cell(r, 10).number_format = '#,##0.00'

            tally_row += 1
            entry_code += 1

        # Part 2 — Trf Entry rows (weekly)
        for last_date, net_val in trf_entries.items():
            r = tally_row
            start_date, end_date = week_ranges[last_date]
            sum_row = pt_row_map.get(last_date, pt_data_start)
            date_str   = last_date.strftime('%d/%m/%y') if pd.notna(last_date) else ''
            tally_date = int(last_date.strftime('%Y%m%d')) if pd.notna(last_date) else ''
            narration  = f"Inv Trf Entries from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"

            for ci in range(1, 11):
                ws_te.cell(r, ci).fill = ORANGE
                ws_te.cell(r, ci).font = BOLD_FONT

            ws_te.cell(r, 1, "")
            ws_te.cell(r, 2, date_str)
            ws_te.cell(r, 3, tally_date)
            ws_te.cell(r, 3).number_format = '0'
            ws_te.cell(r, 4, "Journal")
            ws_te.cell(r, 5, "P- JV- MH")
            ws_te.cell(r, 6, narration)
            ws_te.cell(r, 7, "Cashfree Payments India(Pasfar) Pvt Ltd-V")
            ws_te.cell(r, 8, f"=-J{r}")
            ws_te.cell(r, 8).number_format = '#,##0.00'
            ws_te.cell(r, 9, "Cashfree Payments India(Pasfar) Pvt Ltd-C")
            ws_te.cell(r, 10, f"=Summary!{TRF_COL}{sum_row}")
            ws_te.cell(r, 10).number_format = '#,##0.00'

            tally_row += 1

        # Part 3 — VPA entry rows
        pt_start_date = pt['date_only'].min() if len(pt) > 0 else None
        for ri, row in enumerate(vpa.itertuples(), 0):
            r = tally_row
            vpa_sum_row = vpa_row_map.get(row.date_only, data_start_d + ri)
            date_str   = row.date_only.strftime('%d/%m/%y') if pd.notna(row.date_only) else ''
            tally_date = int(row.date_only.strftime('%Y%m%d')) if pd.notna(row.date_only) else ''
            start_str  = pt_start_date.strftime('%Y-%m-%d') if pt_start_date is not None and pd.notna(pt_start_date) else ''
            end_str    = row.date_only.strftime('%Y-%m-%d') if pd.notna(row.date_only) else ''
            narration  = f"VPA_DETAILS_FROM_VPA{end_str} to {end_str}" if not start_str else f"VPA_DETAILS_FROM_VPA{start_str} to {end_str}"

            for ci in range(1, 11):
                ws_te.cell(r, ci).fill = ORANGE
                ws_te.cell(r, ci).font = BOLD_FONT

            ws_te.cell(r, 1, "")
            ws_te.cell(r, 2, date_str)
            ws_te.cell(r, 3, tally_date)
            ws_te.cell(r, 3).number_format = '0'
            ws_te.cell(r, 4, "Journal")
            ws_te.cell(r, 5, "P- JV- MH")
            ws_te.cell(r, 6, narration)
            ws_te.cell(r, 7, "Cashfree Payments India(Pasfar) Pvt Ltd-V")
            ws_te.cell(r, 8, f"=-J{r}")
            ws_te.cell(r, 8).number_format = '#,##0.00'
            ws_te.cell(r, 9, "Cashfree Payments India(Pasfar) Pvt Ltd-C")
            ws_te.cell(r, 10, f"=Summary!{TRF_COL}{vpa_sum_row}")
            ws_te.cell(r, 10).number_format = '#,##0.00'

            tally_row += 1

        # Column widths for tally entry
        te_widths = {1:12, 2:12, 3:12, 4:10, 5:14, 6:55, 7:42, 8:20, 9:42, 10:20}
        for ci, w in te_widths.items():
            ws_te.column_dimensions[get_column_letter(ci)].width = w

        total_tally_rows = tally_row - 2
        emit(8, TOTAL, "Step 8 complete",
             f"Tally entry: {len(pt)} normal rows + {len(trf_entries)} Trf rows + {len(vpa)} VPA rows = {total_tally_rows} total",
             extra={"tally_rows": total_tally_rows})
    except Exception as e:
        emit_error(f"Step 8 failed: {e}\n{traceback.format_exc()}")
        return

    # ── STEP 9 — Sheet ordering & save ────────────────────────────────────
    emit(9, TOTAL, "Finalising workbook", "Ordering sheets and saving output…")
    try:
        desired = ['transfer report', 'account statement', 'Summary', 'tally entry']
        # Remove any leftover original sheets
        for sname in list(wb.sheetnames):
            if sname not in desired:
                del wb[sname]
        # Reorder
        wb._sheets.sort(key=lambda s: desired.index(s.title) if s.title in desired else 99)
        wb.save(out_path)
        emit(9, TOTAL, "Step 9 complete", f"Workbook saved: {out_path}")
    except Exception as e:
        emit_error(f"Step 9 failed: {e}")
        return

    # ── STEP 10 — Done ────────────────────────────────────────────────────
    emit(10, TOTAL, "Processing complete",
         f"All 4 sheets written: transfer report | account statement | Summary | tally entry. "
         f"Removed {removed} rows | Fixed {suffix_fixed} IDs | {dup_count} duplicates | "
         f"{total_tally_rows} tally rows.",
         done=True,
         extra={
             "removed": removed,
             "suffix_fixed": suffix_fixed,
             "dup_count": dup_count,
             "tally_rows": total_tally_rows,
             "output": out_path,
         })


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(json.dumps({"error": "Usage: cashfree_processor.py <input> <output>"}))
        sys.exit(1)
    process(sys.argv[1], sys.argv[2])
